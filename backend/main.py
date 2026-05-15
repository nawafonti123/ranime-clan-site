from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client, Client
from datetime import datetime
from urllib.parse import urlparse, unquote
import os
import uuid
import csv
from io import BytesIO, StringIO

app = FastAPI(title="RANIME Gaming API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "application-videos")

supabase: Client | None = None
if SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)


def get_supabase() -> Client:
    if supabase is None:
        raise HTTPException(
            status_code=500,
            detail="Supabase env vars are missing: SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY",
        )
    return supabase


def storage_path_from_public_url(video_url: str | None) -> str | None:
    """Extract applications/file.mp4 from a Supabase public storage URL."""
    if not video_url:
        return None

    try:
        parsed = urlparse(video_url)
        marker = f"/storage/v1/object/public/{SUPABASE_BUCKET}/"
        if marker not in parsed.path:
            return None
        path = parsed.path.split(marker, 1)[1]
        return unquote(path) if path else None
    except Exception:
        return None


VIDEO_ALLOWED_TYPES = {
    "video/mp4",
    "video/webm",
    "video/quicktime",
    "video/x-matroska",
}

IMAGE_ALLOWED_TYPES = {
    "image/jpeg",
    "image/png",
    "image/webp",
    "image/gif",
}

CLAN_RANKS = {
    "member": "لاعب عادي",
    "elite": "لاعب نخبة",
    "co_leader": "كو ليدر",
    "leader": "رئيس الكلان",
}

RANK_ORDER_MAP = {"leader": 1, "co_leader": 2, "elite": 3, "member": 4}

EXCEL_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}


def normalize_rank(status: str | None) -> str:
    value = (status or "").strip().lower().replace("-", "_").replace(" ", "_")

    if value in {"leader", "president", "owner", "chief", "رئيس", "رئيس_الكلان"}:
        return "leader"

    if value in {"co_leader", "coleader", "co", "co_leadr", "كو_ليدر", "كوليدر"}:
        return "co_leader"

    if value in {"elite", "النخبة", "نخبة", "لاعب_نخبة"}:
        return "elite"

    return "member"


def row_value(row: dict, possible_names: list[str]) -> str:
    normalized = {
        str(k).strip().lower().replace(" ", "").replace("_", ""): v
        for k, v in row.items()
        if k is not None
    }
    for name in possible_names:
        key = name.strip().lower().replace(" ", "").replace("_", "")
        value = normalized.get(key)
        if value is not None:
            return str(value).strip()
    return ""


def parse_members_file(filename: str, content: bytes) -> list[dict]:
    ext = os.path.splitext(filename or "")[1].lower()

    if ext not in EXCEL_ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="صيغة الملف غير مدعومة. استخدم xlsx أو xlsm أو csv")

    rows: list[dict] = []

    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        rows = [dict(r) for r in reader]
    else:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(status_code=500, detail="مكتبة openpyxl غير مثبتة على السيرفر")

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))

        if not values:
            return []

        headers = [str(v).strip() if v is not None else "" for v in values[0]]
        for values_row in values[1:]:
            item = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                item[header] = values_row[index] if index < len(values_row) else ""
            rows.append(item)

    members: list[dict] = []
    seen: set[str] = set()

    for row in rows:
        player_name = row_value(row, ["Pubg Name", "PUBG Name", "player_name", "Player Name", "Name", "الاسم"])
        status = row_value(row, ["Status", "Rank", "clan_rank", "Clan Rank", "الرتبة", "الحالة"])

        if not player_name:
            continue

        key = player_name.strip().lower()
        if key in seen:
            continue
        seen.add(key)

        clan_rank = normalize_rank(status)
        members.append({
            "player_name": player_name.strip(),
            "pubg_id": "",
            "discord": "",
            "device": "",
            "fps": "",
            "game_role": "",
            "description": "",
            "video_url": None,
            "profile_image_url": None,
            "clan_rank": clan_rank,
            "clan_title": CLAN_RANKS[clan_rank],
            "rank_order": RANK_ORDER_MAP[clan_rank],
            "is_active": True,
            "created_at": datetime.utcnow().isoformat(),
        })

    return members


async def upload_file_to_storage(file: UploadFile, folder: str, allowed_types: set[str], default_name: str) -> tuple[str, str]:
    """Upload any allowed file to Supabase storage and return (public_url, storage_path)."""
    sb = get_supabase()

    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم")

    ext = os.path.splitext(file.filename or default_name)[1] or os.path.splitext(default_name)[1] or ".bin"
    filename = f"{uuid.uuid4().hex}{ext}"
    storage_path = f"{folder}/{filename}"
    content = await file.read()

    try:
        sb.storage.from_(SUPABASE_BUCKET).upload(
            path=storage_path,
            file=content,
            file_options={
                "content-type": file.content_type or "application/octet-stream",
                "upsert": "false",
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Storage upload failed: {e}")

    public_url = sb.storage.from_(SUPABASE_BUCKET).get_public_url(storage_path)
    return public_url, storage_path


async def upload_video_to_storage(video: UploadFile, folder: str) -> tuple[str, str]:
    return await upload_file_to_storage(video, folder, VIDEO_ALLOWED_TYPES, "video.mp4")


async def upload_image_to_storage(image: UploadFile, folder: str) -> tuple[str, str]:
    return await upload_file_to_storage(image, folder, IMAGE_ALLOWED_TYPES, "profile.png")


def safe_remove_storage_file(video_url: str | None) -> None:
    storage_path = storage_path_from_public_url(video_url)
    if not storage_path:
        return
    try:
        get_supabase().storage.from_(SUPABASE_BUCKET).remove([storage_path])
    except Exception:
        pass


@app.get("/")
def root():
    return {"message": "RANIME Gaming Backend Running With Supabase"}


@app.post("/api/apply")
async def apply_to_clan(
    player_name: str = Form(...),
    pubg_id: str = Form(...),
    discord: str = Form(""),
    device: str = Form(""),
    fps: str = Form(""),
    role: str = Form(""),
    description: str = Form(""),
    video: UploadFile = File(...),
    profile_image: UploadFile = File(...),
):
    sb = get_supabase()
    public_url, storage_path = await upload_video_to_storage(video, "applications")
    image_url, image_storage_path = await upload_image_to_storage(profile_image, "profile-images")

    row = {
        "player_name": player_name,
        "pubg_id": pubg_id,
        "discord": discord,
        "device": device,
        "fps": fps,
        "role": role,
        "description": description,
        "video_url": public_url,
        "storage_path": storage_path,
        "profile_image_url": image_url,
        "profile_image_storage_path": image_storage_path,
        "status": "pending",
        "created_at": datetime.utcnow().isoformat(),
    }

    try:
        result = sb.table("applications").insert(row).execute()
        inserted = result.data[0] if result.data else row
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database insert failed: {e}")

    return {
        "success": True,
        "message": "تم رفع طلبك بنجاح، سيتم مراجعته من الإدارة.",
        "application_id": inserted.get("id"),
    }


@app.get("/api/applications")
def get_applications():
    sb = get_supabase()
    result = sb.table("applications").select("*").order("id", desc=True).execute()
    return result.data or []


@app.delete("/api/applications/{application_id}")
def delete_application(application_id: int):
    sb = get_supabase()

    try:
        found = (
            sb.table("applications")
            .select("id, video_url, profile_image_url")
            .eq("id", application_id)
            .limit(1)
            .execute()
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="طلب التقديم غير موجود")

    video_url = found.data[0].get("video_url")
    profile_image_url = found.data[0].get("profile_image_url")
    for file_url in [video_url, profile_image_url]:
        storage_path = storage_path_from_public_url(file_url)
        if storage_path:
            try:
                sb.storage.from_(SUPABASE_BUCKET).remove([storage_path])
            except Exception:
                # لا نوقف حذف الطلب إذا فشل حذف الملف من التخزين
                pass

    try:
        sb.table("applications").delete().eq("id", application_id).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database delete failed: {e}")

    return {"success": True, "message": "تم حذف طلب التقديم بنجاح"}


# =========================
# CLAN MEMBERS MANAGEMENT
# =========================

@app.get("/api/clan-members")
def get_clan_members():
    sb = get_supabase()
    try:
        result = (
            sb.table("clan_members")
            .select("*")
            .eq("is_active", True)
            .order("rank_order", desc=False)
            .order("id", desc=False)
            .execute()
        )
        return result.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Clan members read failed: {e}")




@app.post("/api/clan-members/import")
async def import_clan_members(file: UploadFile = File(...)):
    sb = get_supabase()

    content = await file.read()
    members = parse_members_file(file.filename or "", content)

    if not members:
        raise HTTPException(status_code=400, detail="لم يتم العثور على أعضاء داخل الملف")

    try:
        existing_result = (
            sb.table("clan_members")
            .select("id, player_name")
            .execute()
        )
        existing = existing_result.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Clan members read failed: {e}")

    existing_by_name = {
        (item.get("player_name") or "").strip().lower(): item.get("id")
        for item in existing
        if item.get("player_name")
    }

    inserted_count = 0
    updated_count = 0

    for member in members:
        key = member["player_name"].strip().lower()
        current_id = existing_by_name.get(key)

        try:
            if current_id:
                sb.table("clan_members").update({
                    "clan_rank": member["clan_rank"],
                    "clan_title": member["clan_title"],
                    "rank_order": member["rank_order"],
                    "is_active": True,
                }).eq("id", current_id).execute()
                updated_count += 1
            else:
                result = sb.table("clan_members").insert(member).execute()
                inserted = (result.data or [{}])[0]
                if inserted.get("id"):
                    existing_by_name[key] = inserted.get("id")
                inserted_count += 1
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Import failed for {member['player_name']}: {e}")

    return {
        "success": True,
        "message": f"تم استيراد الأعضاء بنجاح: جديد {inserted_count} / تحديث {updated_count}",
        "imported": inserted_count,
        "updated": updated_count,
        "total": len(members),
    }


@app.post("/api/applications/{application_id}/approve")
def approve_application_as_member(
    application_id: int,
    clan_rank: str = Form("member"),
    custom_title: str = Form(""),
):
    sb = get_supabase()

    if clan_rank not in CLAN_RANKS:
        raise HTTPException(status_code=400, detail="رتبة العضو غير صحيحة")

    rank_order_map = {"leader": 1, "co_leader": 2, "elite": 3, "member": 4}

    try:
        found = sb.table("applications").select("*").eq("id", application_id).limit(1).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Application read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="طلب التقديم غير موجود")

    app_row = found.data[0]
    row = {
        "application_id": application_id,
        "player_name": app_row.get("player_name"),
        "pubg_id": app_row.get("pubg_id"),
        "discord": app_row.get("discord") or "",
        "device": app_row.get("device") or "",
        "fps": app_row.get("fps") or "",
        "game_role": app_row.get("role") or "",
        "description": app_row.get("description") or "",
        "video_url": app_row.get("video_url"),
        "profile_image_url": app_row.get("profile_image_url"),
        "clan_rank": clan_rank,
        "clan_title": custom_title or CLAN_RANKS[clan_rank],
        "rank_order": rank_order_map[clan_rank],
        "is_active": True,
        "created_at": datetime.utcnow().isoformat(),
    }

    try:
        inserted = sb.table("clan_members").insert(row).execute()
        sb.table("applications").update({"status": "approved", "reviewed_at": datetime.utcnow().isoformat()}).eq("id", application_id).execute()
        return {"success": True, "message": "تم قبول اللاعب وإضافته إلى أعضاء الكلان", "member": (inserted.data or [row])[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Approve application failed: {e}")


@app.put("/api/clan-members/{member_id}")
def update_clan_member(
    member_id: int,
    clan_rank: str = Form("member"),
    custom_title: str = Form(""),
    is_active: bool = Form(True),
):
    sb = get_supabase()

    if clan_rank not in CLAN_RANKS:
        raise HTTPException(status_code=400, detail="رتبة العضو غير صحيحة")

    rank_order_map = {"leader": 1, "co_leader": 2, "elite": 3, "member": 4}
    updates = {
        "clan_rank": clan_rank,
        "clan_title": custom_title or CLAN_RANKS[clan_rank],
        "rank_order": rank_order_map[clan_rank],
        "is_active": is_active,
    }

    try:
        result = sb.table("clan_members").update(updates).eq("id", member_id).execute()
        return {"success": True, "message": "تم تحديث عضو الكلان", "member": (result.data or [updates])[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Clan member update failed: {e}")


@app.delete("/api/clan-members/{member_id}")
def delete_clan_member(member_id: int):
    sb = get_supabase()
    try:
        sb.table("clan_members").update({"is_active": False}).eq("id", member_id).execute()
        return {"success": True, "message": "تم إزالة العضو من صفحة الكلان"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Clan member delete failed: {e}")


# =========================
# SITE VIDEOS MANAGEMENT
# =========================

@app.get("/api/site-videos")
def get_site_videos():
    sb = get_supabase()
    try:
        result = (
            sb.table("site_videos")
            .select("*")
            .eq("is_active", True)
            .order("slot", desc=False)
            .order("id", desc=False)
            .execute()
        )
        return result.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Videos read failed: {e}")


@app.post("/api/site-videos")
async def create_site_video(
    title: str = Form(...),
    description: str = Form(""),
    slot: int = Form(1),
    video: UploadFile = File(...),
):
    sb = get_supabase()
    public_url, storage_path = await upload_video_to_storage(video, "site-videos")

    row = {
        "title": title,
        "description": description,
        "slot": slot,
        "video_url": public_url,
        "storage_path": storage_path,
        "is_active": True,
        "created_at": datetime.utcnow().isoformat(),
    }

    try:
        result = sb.table("site_videos").insert(row).execute()
        return {"success": True, "message": "تم إضافة الفيديو بنجاح", "video": (result.data or [row])[0]}
    except Exception as e:
        safe_remove_storage_file(public_url)
        raise HTTPException(status_code=500, detail=f"Video insert failed: {e}")


@app.put("/api/site-videos/{video_id}")
async def update_site_video(
    video_id: int,
    title: str = Form(...),
    description: str = Form(""),
    slot: int = Form(1),
    video: UploadFile | None = File(None),
):
    sb = get_supabase()

    try:
        found = sb.table("site_videos").select("*").eq("id", video_id).limit(1).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="الفيديو غير موجود")

    updates = {
        "title": title,
        "description": description,
        "slot": slot,
    }

    old_video_url = found.data[0].get("video_url")

    if video is not None:
        public_url, storage_path = await upload_video_to_storage(video, "site-videos")
        updates["video_url"] = public_url
        updates["storage_path"] = storage_path

    try:
        result = sb.table("site_videos").update(updates).eq("id", video_id).execute()
        if video is not None:
            safe_remove_storage_file(old_video_url)
        return {"success": True, "message": "تم تحديث الفيديو", "video": (result.data or [updates])[0]}
    except Exception as e:
        if video is not None:
            safe_remove_storage_file(updates.get("video_url"))
        raise HTTPException(status_code=500, detail=f"Video update failed: {e}")


@app.delete("/api/site-videos/{video_id}")
def delete_site_video(video_id: int):
    sb = get_supabase()

    try:
        found = sb.table("site_videos").select("id, video_url").eq("id", video_id).limit(1).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="الفيديو غير موجود")

    safe_remove_storage_file(found.data[0].get("video_url"))

    try:
        sb.table("site_videos").delete().eq("id", video_id).execute()
        return {"success": True, "message": "تم حذف الفيديو"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video delete failed: {e}")


# =========================
# VISITOR VIDEO REQUESTS
# =========================

@app.post("/api/video-requests")
async def create_video_request(
    visitor_name: str = Form(...),
    contact: str = Form(""),
    title: str = Form(...),
    description: str = Form(""),
    video: UploadFile = File(...),
):
    sb = get_supabase()
    public_url, storage_path = await upload_video_to_storage(video, "video-requests")

    row = {
        "visitor_name": visitor_name,
        "contact": contact,
        "title": title,
        "description": description,
        "video_url": public_url,
        "storage_path": storage_path,
        "status": "pending",
        "created_at": datetime.utcnow().isoformat(),
    }

    try:
        result = sb.table("video_requests").insert(row).execute()
        return {"success": True, "message": "تم إرسال طلب الفيديو للإدارة", "request": (result.data or [row])[0]}
    except Exception as e:
        safe_remove_storage_file(public_url)
        raise HTTPException(status_code=500, detail=f"Video request insert failed: {e}")


@app.get("/api/video-requests")
def get_video_requests(status: str | None = None):
    sb = get_supabase()

    try:
        query = sb.table("video_requests").select("*")
        if status:
            query = query.eq("status", status)
        result = query.order("id", desc=True).execute()
        return result.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video requests read failed: {e}")


@app.post("/api/video-requests/{request_id}/approve")
def approve_video_request(request_id: int):
    sb = get_supabase()

    try:
        found = sb.table("video_requests").select("*").eq("id", request_id).limit(1).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video request read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="طلب الفيديو غير موجود")

    req = found.data[0]

    try:
        sb.table("video_requests").update(
            {"status": "approved", "reviewed_at": datetime.utcnow().isoformat()}
        ).eq("id", request_id).execute()

        design_row = {
            "title": req.get("title") or "تصميم جديد",
            "description": req.get("description") or "",
            "slot": 99,
            "video_url": req.get("video_url"),
            "storage_path": req.get("storage_path"),
            "is_active": True,
            "created_at": datetime.utcnow().isoformat(),
        }
        inserted = sb.table("site_videos").insert(design_row).execute()
        return {"success": True, "message": "تم قبول الفيديو ونشره في قسم التصاميم", "video": (inserted.data or [design_row])[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video request approve failed: {e}")


@app.post("/api/video-requests/{request_id}/reject")
def reject_video_request(request_id: int):
    sb = get_supabase()

    try:
        found = sb.table("video_requests").select("*").eq("id", request_id).limit(1).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video request read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="طلب الفيديو غير موجود")

    safe_remove_storage_file(found.data[0].get("video_url"))

    try:
        sb.table("video_requests").update(
            {"status": "rejected", "reviewed_at": datetime.utcnow().isoformat()}
        ).eq("id", request_id).execute()
        return {"success": True, "message": "تم رفض طلب الفيديو"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video request reject failed: {e}")


@app.delete("/api/video-requests/{request_id}")
def delete_video_request(request_id: int):
    sb = get_supabase()

    try:
        found = sb.table("video_requests").select("*").eq("id", request_id).limit(1).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video request read failed: {e}")

    if not found.data:
        raise HTTPException(status_code=404, detail="طلب الفيديو غير موجود")

    req = found.data[0]
    status = req.get("status")

    # إذا كان الطلب منشوراً، لا نحذف ملف التخزين هنا لأن التصميم المنشور في site_videos يستخدم نفس الرابط.
    if status != "approved":
        safe_remove_storage_file(req.get("video_url"))

    try:
        sb.table("video_requests").delete().eq("id", request_id).execute()
        return {"success": True, "message": "تم حذف طلب التصميم"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Video request delete failed: {e}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
